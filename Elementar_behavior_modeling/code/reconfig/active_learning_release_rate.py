"""Release active learning: selects experiments, 
runs experiments, and updates behavior prediction network online.

Reads/writes npy/pth/xlsx in current directory; requires GX camera and S826 coil hardware.
"""
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import torch.nn as nn
import torch.optim as optim
from scipy.stats import multivariate_normal
from torch.utils.data import DataLoader, TensorDataset
import sys
import numpy as np
import cv2
import pickle
import os
import time
import S826
import threading
import torch
import openpyxl
from openpyxl.styles import Font, Alignment
import pandas as pd
from utils import moving_average, plot_monitoring_curves
global model
device = 'cuda' if torch.cuda.is_available else 'cpu'
print('\n\nDevice Used:', device)

_CODE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_DEVICES_DIR = os.path.join(_CODE_DIR, "devices")
if _DEVICES_DIR not in sys.path:
    sys.path.insert(0, _DEVICES_DIR)

import gxipy as gx


input_dim = 4
# Discrete grid: flux x frequency x pitch x direction
grid_shape = (11, 21, 37, 73)

probability_scale = 20
kernel_size = 5
confidence_scale = 0.1

if os.path.exists('confidence_space.npy'):
    confidence_space = np.load('confidence_space.npy')
else:
    confidence_space = np.zeros(grid_shape)

if os.path.exists('steps.npy'):
    steps = np.load('steps.npy')
else:
    steps = [0]

if os.path.exists('min_confidences.npy'):
    min_confidences = np.load('min_confidences.npy')
    avg_confidences = np.load('avg_confidences.npy')
else:
    min_confidences = [0]
    avg_confidences = [0]

if os.path.exists('inputs.npy'):
    inputs = np.load('inputs.npy')
    output_pred = np.load('output_pred.npy')
    output_exp = np.load('output_exp.npy')
    errors = np.load('errors.npy')
    error_rates = np.load('error_rates.npy')
else:
    inputs = np.zeros(4)
    output_pred = [0]
    output_exp = [0]
    errors = [0]
    error_rates = [0]

# Grid coordinates for multivariate Gaussian kernel in update_confidence_space
X, Y, Z, W = np.meshgrid(
    np.linspace(0, grid_shape[0], grid_shape[0]),
    np.linspace(0, grid_shape[1], grid_shape[1]),
    np.linspace(0, grid_shape[2], grid_shape[2]),
    np.linspace(0, grid_shape[3], grid_shape[3]),
    indexing='ij'
)

pos = np.empty(X.shape + (input_dim,))
pos[:, :, :, :, 0] = X
pos[:, :, :, :, 1] = Y
pos[:, :, :, :, 2] = Z
pos[:, :, :, :, 3] = W

def write_to_excel(self, flux, frequency, pitch, direction, lr, elapsed_times, gray_scale, file_order):
    """Append one trajectory row to data_{order}.xlsx for this experiment."""
    filename = f'data_{file_order}.xlsx'
    if os.path.exists(filename):
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        worksheet['A1'] = 'Flux'
        worksheet['B1'] = 'Frequency'
        worksheet['C1'] = 'Pitch'
        worksheet['D1'] = 'Direction'
        worksheet['E1'] = 'lr'
        worksheet['F1'] = 'Elapsed Time'
        worksheet['G1'] = 'Gray Scale'

        for col in range(1, 8):
            worksheet.cell(row=1, column=col).font = Font(bold=True)
            worksheet.cell(row=1, column=col).alignment = Alignment(horizontal='center', vertical='center')

    last_row = worksheet.max_row + 1

    for i in range(len(elapsed_times)):
        worksheet.cell(row=last_row, column=1, value=flux)
        worksheet.cell(row=last_row, column=2, value=frequency)
        worksheet.cell(row=last_row, column=3, value=pitch)
        worksheet.cell(row=last_row, column=4, value=direction)
        worksheet.cell(row=last_row, column=5, value=lr)
        worksheet.cell(row=last_row, column=6, value=elapsed_times[i])
        worksheet.cell(row=last_row, column=7, value=gray_scale[i])
        last_row += 1

    workbook.save(filename)


class Camera:
    """Industrial camera acquisition + ROI gray scale statistics; run completes one acquisition sequence under set magnetic field"""

    def __init__(self):
        self.images_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'new_images')

    def go_left(self):
        coil.update_field(19, 49, 61, 210)

    def go_right(self):
        coil.update_field(18, 53, 19, 26)

    def recollect(self):
        coil.update_field(5, 5, 90, 0)

    def get_image(self):
        raw_image = self.cam.data_stream[0].get_image()
        if raw_image is None:
            print("Getting image failed.")

        rgb_image = raw_image.convert("RGB")
        rgb_image.image_improvement(self.color_correction_param, self.contrast_lut, self.gamma_lut)
        numpy_image = rgb_image.get_numpy_array()
        frame = cv2.cvtColor(np.asarray(numpy_image), cv2.COLOR_BGR2RGB)

        height, width, _ = frame.shape
        roi = frame[:, :, :]
        gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
        sorted_pixels = np.sort(gray, axis=None)
        low_index = int(sorted_pixels.size * 0.5)
        high_index = int(sorted_pixels.size * 0.9)
        gray_min_mean = np.mean(sorted_pixels[low_index:high_index])
        font = cv2.FONT_HERSHEY_SIMPLEX
        font_scale = 0.5
        font_color = (255, 0, 0)
        line_type = 2

        cv2.putText(frame, f"gray_scale:{gray_min_mean}", (int(height/2), int(width/2)), font, font_scale, font_color, line_type)

        cv2.imshow("Image", frame)
        cv2.waitKey(1)

        return frame, gray_min_mean

    def run(self, flux, frequency, pitch, direction, lr, data_file_order):
        self.data_file_order = data_file_order
        self.images_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'new_images')
        device_manager = gx.DeviceManager()
        dev_num, dev_info_list = device_manager.update_device_list()
        if dev_num == 0:
            print("Number of enumerated devices is 0")
            return

        self.cam = device_manager.open_device_by_index(1)

        self.cam.Width.set(600)
        self.cam.Height.set(190)
        self.cam.OffsetX.set(1504)
        self.cam.OffsetY.set(650)

        self.cam.TriggerMode.set(gx.GxSwitchEntry.OFF)
        self.cam.ExposureTime.set(2500)
        self.cam.BalanceWhiteAuto.set(1)
        self.cam.Gain.set(0)

        self.cam.UserSetSelector.set(1)
        self.cam.UserSetSave.send_command()
        if self.cam.GammaParam.is_readable():
            gamma_value = self.cam.GammaParam.get()
            self.gamma_lut = gx.Utility.get_gamma_lut(gamma_value)
        else:
            self.gamma_lut = None
        if self.cam.ContrastParam.is_readable():
            contrast_value = self.cam.ContrastParam.get()
            self.contrast_lut = gx.Utility.get_contrast_lut(contrast_value)
        else:
            self.contrast_lut = None
        if self.cam.ColorCorrectionParam.is_readable():
            self.color_correction_param = self.cam.ColorCorrectionParam.get()
        else:
            self.color_correction_param = 0

        self.cam.data_stream[0].set_acquisition_buffer_number(2)
        self.cam.stream_on()

        self.init_time = time.time()
        last_time = 0
        self.flux = flux
        self.frequency = frequency
        self.pitch = pitch
        self.direction = direction
        self.lr = lr

        result, gray_mean = self.get_image()
        self.start_time = time.time()
        elapsed_time = time.time() - self.start_time

        coil.update_field(self.flux, self.frequency, self.pitch, self.direction)

        gray_means = np.array(gray_mean)
        elapsed_times = elapsed_time

        while elapsed_time < 4.1:
            result, gray_mean = self.get_image()

            if int(10*elapsed_time) - int(10*last_time) > 0:
                gray_means = np.append(gray_means,  np.array(gray_mean))
                elapsed_times = np.append(elapsed_times, elapsed_time)

            last_time = elapsed_time
            elapsed_time = time.time() - self.start_time

        while elapsed_time < 15:
            self.recollect()
            elapsed_time = time.time() - self.start_time

        write_to_excel(self, self.flux, self.frequency, self.pitch, self.direction, self.lr, elapsed_times, gray_means, self.data_file_order)

        coil.stop()
        cv2.destroyAllWindows()

    def calculate_release_gray_scale(self, flux, frequency, pitch, direction):
        """Read latest experiment for this parameter set from current xlsx, take final gray scale at lr=0; write summary to true_gray_scales.xlsx."""
        data = pd.read_excel(f'data_{self.data_file_order}.xlsx')

        relevant_data_df = pd.DataFrame()

        for idx in range(len(data) - 1, -1, -1):
            row = data.iloc[idx]

            if (row['Flux'] == flux and
                    row['Frequency'] == frequency and
                    row['Pitch'] == pitch and
                    row['Direction'] == direction):
                relevant_data_df = pd.concat([relevant_data_df, row.to_frame().T], ignore_index=True)
            else:
                break

        if relevant_data_df.empty:
            print("No matching experiment data found.")
            return None

        lr_data = relevant_data_df[relevant_data_df['lr'] == 0].copy()

        final_elapsed_time = lr_data['Elapsed Time'].max()
        final_gray_scale = lr_data.loc[lr_data['Elapsed Time'] == final_elapsed_time, 'Gray Scale'].values[0]

        save_data = {
            'Flux': flux,
            'Frequency': frequency,
            'Pitch': pitch,
            'Direction': direction,
            'Final_Gray': final_gray_scale
        }

        try:
            existing_data = pd.read_excel('true_gray_scales.xlsx')
        except FileNotFoundError:
            existing_data = pd.DataFrame(columns=['Flux', 'Frequency', 'Pitch', 'Direction', 'Velocity'])

        new_data = pd.DataFrame([save_data])
        new_data = pd.concat([existing_data, new_data], ignore_index=True)

        new_data.to_excel('true_gray_scales.xlsx', index=False)

        return final_gray_scale


class Coil:
    """S826 board drives rotating magnetic field."""

    def __init__(self):
        self.coil = S826.S826()
        self.flux = 0
        self.frequency = 0
        self.pitch = 0
        self.direction = 0

        self.state = 1
        self.start_time = time.time()

    def rotating_field(self):
        while self.state:
            self.rotating_field_signal()

    def rotating_field_signal(self):
        self.coil.rotating_field(flux_density=self.flux, frequency=self.frequency, pitch=self.pitch, direction=self.direction)

    def update_field(self, new_flux, new_freq, new_pitch, new_direct):
        self.flux = new_flux
        self.frequency = new_freq
        self.pitch = new_pitch
        self.direction = new_direct

    def stop(self):
        self.update_field(0, 0, 0, 0)
        time.sleep(0.1)
        self.state = 0

class BehaviorNNOnline(nn.Module):
    """Lightweight NN for online behavior prediction (field params -> gray scale)."""

    def __init__(self, input_dim=4):
        super(BehaviorNNOnline, self).__init__()
        self.fc1 = nn.Linear(input_dim, 64)
        self.fc2 = nn.Linear(64, 128)
        self.fc3 = nn.Linear(128, 64)
        self.fc4 = nn.Linear(64, 1)

    def forward(self, x):
        x1 = nn.ReLU()(self.fc1(x))
        x2 = nn.ReLU()(self.fc2(x1))
        x3 = nn.ReLU()(self.fc3(x2))
        output = self.fc4(x3)
        return output

model = BehaviorNNOnline()
if os.path.exists('robot_control_model.pth'):
    model.load_state_dict(torch.load('robot_control_model.pth'))
    model.eval()
else:
    torch.save(model.state_dict(), 'robot_control_model.pth')

optimizer = optim.Adam(model.parameters(), lr=0.001)
criterion = nn.MSELoss()

def train_model(X, y):
    for i in range(100):
        model.train()
        optimizer.zero_grad()
        predictions = model(X)
        loss = criterion(predictions, y)
        loss.backward()
        optimizer.step()

def select_next_experiment(confidence_space):
    """Sample one point on grid via softmax(exp(-confidence * scale))."""
    probabilities = np.exp(-confidence_space * probability_scale)
    probabilities /= probabilities.sum()
    index = np.random.choice(np.arange(confidence_space.size), p=probabilities.flatten())
    selected_position = np.unravel_index(index, confidence_space.shape)

    selected_confidence = confidence_space[selected_position]

    min_index = np.unravel_index(np.argmin(confidence_space), confidence_space.shape)
    min_confidence = confidence_space[min_index]
    return selected_position, min_confidence


def update_confidence_space(location, bias_value, kernel_size=2):
    global confidence_space, X, Y, Z, W, pos

    rv = multivariate_normal(mean=location,
                             cov=kernel_size ** 2 * np.eye(input_dim))
    gaussian = rv.pdf(pos)

    if np.max(gaussian) > 0:
        gaussian = gaussian / np.max(gaussian) * bias_value
        confidence_space += gaussian
    else:
        print("Warning: Gaussian values are too small or zero.")
        return


camera = Camera()
plt.figure(figsize=(10, 6))

# Outer loop: xlsx file volume index; inner loop: iterations per volume
for data_file_order in range(200):
    for iteration in range(10):
        current_experiment, min_confidence = select_next_experiment(confidence_space)
        avg_confidence = np.mean(confidence_space)

        steps = np.append(steps, np.max(steps)+1)
        min_confidences = np.append(min_confidences, min_confidence)
        avg_confidences = np.append(avg_confidences, avg_confidence)

        X_new = np.array(current_experiment)
        # Grid index -> [0,1]^4, then scale to physical range for printing
        X_new_scaled = (X_new - np.array([0, 0, 0, 0])) / np.array([10, 20, 36, 72])
        X_new_human = X_new_scaled * np.array([20, 40, 180, 360])
        print('X_new_human', X_new_human)

        inputs = np.append(inputs, X_new_scaled)
        X_new_scaled = torch.FloatTensor(X_new_scaled)
        y_new = model(X_new_scaled).detach().numpy()
        output_pred = np.append(output_pred, y_new)

        flux, freq, p, d = 2*X_new[0], 2*X_new[1], 5*X_new[2], 5*X_new[3]

        lrs = [0]
        for lr in lrs:
            cv2.waitKey(10)
            coil = Coil()
            cv2.waitKey(10)

            camera_thread = threading.Thread(target=camera.run, args=(flux, freq, p, d, lr, data_file_order))
            coil_thread = threading.Thread(target=coil.rotating_field)

            camera_thread.start()
            coil_thread.start()

            camera_thread.join()
            coil_thread.join()

            cv2.waitKey(10)
            coil.stop()
            cv2.waitKey(100)
            coil.stop()

        experimental_gray = camera.calculate_release_gray_scale(flux, freq, p, d)
        output_exp = np.append(output_exp, experimental_gray)

        print('y_real:', experimental_gray)
        print('y_predict:', y_new)
        error = np.abs((experimental_gray - y_new))
        error_rate = np.abs((experimental_gray - y_new) / experimental_gray)

        errors = np.append(errors, error)
        error_rates = np.append(error_rates, error_rate)

        print('error', error)
        print('error_rate', error_rate)

        confidence_value = 1 - np.tanh(error * error_rate * confidence_scale)
        print('confidence_value', confidence_value)
        update_confidence_space(current_experiment, confidence_value, kernel_size=kernel_size)

        new_X_scaled_np = np.array(X_new_scaled)
        new_Y_np = np.array(experimental_gray)

        try:
            dataset = torch.load('dataset.pth')
            existing_X = dataset.tensors[0]
            existing_y = dataset.tensors[1]
        except FileNotFoundError:
            existing_X = torch.FloatTensor()
            existing_y = torch.FloatTensor(  )

        new_X = torch.FloatTensor(new_X_scaled_np).view(-1, input_dim)
        new_y = torch.FloatTensor(new_Y_np).view(-1, 1)

        updated_X = torch.cat((existing_X, new_X), dim=0)
        updated_y = torch.cat((existing_y, new_y), dim=0)

        dataset = TensorDataset(updated_X, updated_y)

        torch.save(dataset, 'dataset.pth')

        batch_size = 128
        dataloader = DataLoader(dataset, batch_size=batch_size, shuffle=True)

        for inputs, targets in dataloader:
            train_model(inputs, targets)

        torch.save(model.state_dict(), 'robot_control_model.pth')

        np.save('inputs.npy', inputs)
        np.save('output_pred', output_pred)
        np.save('output_exp', output_exp)
        np.save('confidence_space.npy', confidence_space)
        np.save('steps.npy', steps)
        np.save('min_confidences.npy', min_confidences)
        np.save('avg_confidences.npy', avg_confidences)
        np.save('errors.npy', errors)
        np.save('error_rates.npy', error_rates)

        print('Min confidence:', min_confidence, 'Average confidence:', avg_confidence)

        # Generate monitoring plots
        plot_monitoring_curves(steps, min_confidences, avg_confidences, errors, error_rates)